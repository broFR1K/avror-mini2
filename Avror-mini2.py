from colorama import init, Fore
import time
import os
import requests
import json
import socket
import whois
from concurrent.futures import ThreadPoolExecutor
import art
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm
import random
from datetime import datetime

# Инициализация colorama для цветного вывода
init(autoreset=True)

# Функция для очистки консоли в зависимости от ОС
def clear_console():
    if os.name == 'nt':
        os.system('cls')
    else:
        os.system('clear')

# Приветственное сообщение
def print_welcome_message():
    print(Fore.RED + '░█████╗░██╗░░░██╗██████╗░░█████╗░██████╗░')
    print(Fore.RED + '██╔══██╗██║░░░██║██╔══██╗██╔══██╗██╔══██╗')
    print(Fore.RED + '███████║╚██╗░██╔╝██████╔╝██║░░██║██████╔╝')
    print(Fore.RED + '██╔══██║░╚████╔╝░██╔══██╗██║░░██║██╔══██╗')
    print(Fore.RED + '██║░░██║░░╚██╔╝░░██║░░██║╚█████╔╝██║░░██║')
    print(Fore.RED + '╚═╝░░╚═╝░░░╚═╝░░░╚═╝░░╚═╝░╚════╝░╚═╝░░╚═╝')
    print(Fore.LIGHTYELLOW_EX + "Version : AVROR-mini")
    print(" ")
    print(Fore.CYAN + '> Разработчик : broFR1k')
    print(Fore.CYAN + "> Канал telegram : https://t.me/AVROR_S'")
    print(Fore.CYAN + ">Discord Server :   https://t.me/AVROR_S/73")
    print(Fore.CYAN + "> Версия : BETTA 2.0")
    print(" ")
    print(Fore.GREEN + "=============================================================")
    print(Fore.GREEN + "Данный Софт Создан Исключительно для ознакомительных целей")
    print(Fore.GREEN + "Разработчик не несет ответственность за ваши действия")
    print(Fore.GREEN + "=============================================================")
    time.sleep(2)
    print(" ")
    print(Fore.RED + "[1] Информация об IP")
    print(Fore.RED + "[2] Информация об Тел. Номер")
    print(Fore.RED + "[3] Пробив по нику")
    print(Fore.RED + "[4] Пробив по домену")
    print(Fore.RED + "[5] Пробив по тг [БД]")
    print(Fore.RED + "[6] Пробив по номера [БД]")
    print(" ")
# Основное меню
def main():
    while True:
        clear_console()
        print_welcome_message()

        try:
            user_choice = input("Выберите опцию (1-6) или 'q' для выхода: ").strip()

            clear_console()

            if user_choice == "1":
                art.tprint("AVROR_IP")
                tracker = IPTracker()
                tracker.track_ip()

            elif user_choice == "2":
                art.tprint("AVROR_Phone")
                phone_info = PhoneNumberInfo()
                phone_info.print_number_info()

            elif user_choice == "3":
                art.tprint("AVROR_Name")
                username = input("Введите ник: ").strip()
                tracker = UsernameTracker(username)
                tracker.track()
                tracker.show_results()

            elif user_choice == "4":
                art.tprint("AVROR_Domen")
                domain = input(Fore.YELLOW + "Введите доменное имя сайта (например, example.com): ").strip()
                scanner = DomainScanner(domain)
                scanner.run()

            elif user_choice == "5":
             folder_path = "Telegram"  # Укажите путь к папке с файлами
             os.system('cls')
             art.tprint("AVROR_TG-BD")
             search_term = input("Введите ID: ")  # Ввод числа

             searcher = ExcelSearchertg(folder_path)
             searcher.search_in_folder(search_term)


            elif user_choice == "6":
             folder_path = "Phone"  # Укажите путь к папке с файлами
             os.system('cls')
             art.tprint("AVROR_PH-BD")
             search_term = input("Введите номер без '+' или с '+': ")  # Ввод числа

             searcher = ExcelSearcherphone(folder_path)
             searcher.search_in_folder(search_term)


            elif user_choice.lower() == 'q':
                print(Fore.GREEN + "Выход из программы. До свидания!")
                break

            else:
                print(Fore.RED + "Выберите 1-6 или нажмите 'q' чтобы выйти.")

            input(Fore.GREEN + "\nНажмите 'Enter' для продолжения...")

        except Exception as e:
            print(Fore.RED + f"Произошла ошибка: {str(e)}")
            input(Fore.GREEN + "\nНажмите 'Enter' для продолжения...")
#__________________________________________________________________________________________________________#
# Класс для работы с IP
class IPTracker:
    def __init__(self):
        self.ip = None
        self.ip_data = None
    
    def get_ip(self):
        self.ip = input("\n Введите IP адрес (или 'Enter' для выхода из программы): ").strip()
        if not self.ip:
            return False  # Возврат в главное меню, если IP не введен
        return True

    def fetch_ip_data(self):
        if not self.ip:
            print("Данные об IP-адресе недоступны.")
            return
        try:
            req_api = requests.get(f"http://ipwho.is/{self.ip}")
            self.ip_data = req_api.json()
        except requests.RequestException as e:
            print("Ошибка в получении данных IP-адреса:", e)
    
    def display_ip_info(self):
        if not self.ip_data:
            print("Данные об IP-адресе недоступны.")
            return
        print('\n           🄰🅅🅁🄾🅁 🄸🄿                ')
        print("\n IP Жертвы           :", self.ip)
        print(" Тип IP              :", self.ip_data.get("type"))
        print(" Страна              :", self.ip_data.get("country"))
        print(" Код страны          :", self.ip_data.get("country_code"))
        print(" Город               :", self.ip_data.get("city"))
        print(" Континент           :", self.ip_data.get("continent"))
        print(" Код континента      :", self.ip_data.get("continent_code"))
        print(" Регион              :", self.ip_data.get("region"))
        print(" Код региона         :", self.ip_data.get("region_code"))
        print(" Широта              :", self.ip_data.get("latitude"))
        print(" Долгота             :", self.ip_data.get("longitude"))
        
        lat = self.ip_data.get('latitude')
        lon = self.ip_data.get('longitude')
        print(" Карта               :", f"https://www.google.com/maps/@{lat},{lon},8z")
        print(" Евросоюз            :", self.ip_data.get("is_eu"))
        print(" Почтовый индекс     :", self.ip_data.get("postal"))
        print(" Код вызова          :", self.ip_data.get("calling_code"))
        print(" Капитал             :", self.ip_data.get("capital"))
        print(" Границы             :", self.ip_data.get("borders"))
        print(" Флаг Страны         :", self.ip_data["flag"]["emoji"])
        print(" ASN                 :", self.ip_data["connection"]["asn"])
        print(" ORG                 :", self.ip_data["connection"]["org"])
        print(" Интернет провайдер  :", self.ip_data["connection"]["isp"])
        print(" Домен               :", self.ip_data["connection"]["domain"])
        print(" ID                  :", self.ip_data["timezone"]["id"])
        print(" Сокращение          :", self.ip_data["timezone"]["abbr"])
        print(" Летнее время        :", self.ip_data["timezone"]["is_dst"])
        print(" Offset              :", self.ip_data["timezone"]["offset"])
        print(" UTC                 :", self.ip_data["timezone"]["utc"])
        print(" Текущее время       :", self.ip_data["timezone"]["current_time"])
    
    def track_ip(self):
        while True:
            if not self.get_ip():
                print("Возврат в главное меню...")
                break  # Выход из цикла и возврат в главное меню
            self.fetch_ip_data()
            time.sleep(2)
            self.display_ip_info()

# Класс для работы с доменом
class DomainScanner:
    def __init__(self, domain, start_port=1, end_port=8090):
        self.domain = domain
        self.start_port = start_port
        self.end_port = end_port
        self.ip_address = None
        self.domain_info = None

    def get_ip_address(self):
        try:
            self.ip_address = socket.gethostbyname(self.domain)
            return self.ip_address
        except socket.gaierror:
            print(Fore.RED + f"Не удалось получить IP-адрес для домена: {self.domain}")
            return None

    def get_domain_info(self):
        try:
            self.domain_info = whois.whois(self.domain)
            return self.domain_info
        except Exception as e:
            print(Fore.RED + f"Ошибка при получении информации о домене: {e}")
            return None

    def display_domain_info(self):
        if self.domain_info:
            print(Fore.GREEN + "\nИнформация о домене:")
            print(Fore.GREEN + f"Доменное имя: {self.domain_info.domain_name}")
            print(Fore.GREEN + f"Дата регистрации: {self.domain_info.creation_date}")
            print(Fore.GREEN + f"Дата обновления: {self.domain_info.updated_date}")
            print(Fore.GREEN + f"Дата окончания: {self.domain_info.expiration_date}")
            print(Fore.GREEN + f"Регистратор: {self.domain_info.registrar}")
            print(Fore.GREEN + f"Страна: {self.domain_info.country}")

            print(Fore.GREEN + f"Статус домена: {self.domain_info.status}")
            print(Fore.GREEN + f"DNS-серверы: {', '.join(self.domain_info.name_servers) if self.domain_info.name_servers else 'Нет данных'}")

            print(Fore.GREEN + f"Email администратора: {self.domain_info.emails if self.domain_info.emails else 'Нет данных'}")
            print(Fore.GREEN + f"Телефон администратора: {self.domain_info.phone if self.domain_info.phone else 'Нет данных'}")
            print(Fore.GREEN + f"Имя администратора: {self.domain_info.admin if self.domain_info.admin else 'Нет данных'}")
        else:
            print(Fore.RED + "Информация о домене недоступна.")

    def scan_port(self, port):
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
                sock.settimeout(1)  # Тайм-аут в 1 секунду
                result = sock.connect_ex((self.ip_address, port))
                return port if result == 0 else None
        except Exception as e:
            print(Fore.RED + f"Ошибка при сканировании порта {port}: {e}")
            return None

    def scan_ports(self):
        if not self.ip_address:
            print(Fore.RED + "IP-адрес не найден. Сначала получите IP-адрес.")
            return []

        open_ports = []
        with ThreadPoolExecutor(max_workers=100) as executor:
            futures = {executor.submit(self.scan_port, port): port for port in range(self.start_port, self.end_port + 1)}
            for future in futures:
                try:
                    port = future.result()
                    if port is not None:
                        open_ports.append(port)
                except Exception as e:
                    print(Fore.RED + f"Ошибка в потоке сканирования порта: {e}")
        return open_ports

    def run(self):
        # Получение информации о домене
        self.get_domain_info()
        self.display_domain_info()

        # Получение IP-адреса
        ip_address = self.get_ip_address()
        if ip_address:
            print(Fore.CYAN + f"\nIP-адрес сайта {self.domain}: {ip_address}")

            # Сканирование портов
            open_ports = self.scan_ports()
            if open_ports:
                print(Fore.GREEN + f"\nОткрытые порты на {ip_address}: {', '.join(map(str, open_ports))}")
            else:
                print(Fore.YELLOW + f"\nНет открытых портов на {ip_address} в диапазоне {self.start_port}-{self.end_port}.")
        else:
            print(Fore.RED + "Сканирование портов невозможно, так как IP-адрес не найден.")

# Класс для работы с пользователем
class UsernameTracker:
    def __init__(self, username):
        self.username = username
        self.results = {}
        self.social_media = [
            {"url": "https://www.github.com/{}", "name": "GitHub"},
            {"url": "https://www.pinterest.com/{}", "name": "Pinterest"},
            {"url": "https://www.tumblr.com/{}", "name": "Tumblr"},
            {"url": "https://www.youtube.com/{}", "name": "Youtube"},
            {"url": "https://soundcloud.com/{}", "name": "SoundCloud"},
            {"url": "https://www.snapchat.com/add/{}", "name": "Snapchat"},
            {"url": "https://www.tiktok.com/@{}", "name": "TikTok"},
            {"url": "https://www.behance.net/{}", "name": "Behance"},
            {"url": "https://www.medium.com/@{}", "name": "Medium"},
            {"url": "https://www.quora.com/profile/{}", "name": "Quora"},
            {"url": "https://www.flickr.com/people/{}", "name": "Flickr"},
            {"url": "https://www.periscope.tv/{}", "name": "Periscope"},
            {"url": "https://www.twitch.tv/{}", "name": "Twitch"},
            {"url": "https://www.dribbble.com/{}", "name": "Dribbble"},
            {"url": "https://www.stumbleupon.com/stumbler/{}", "name": "StumbleUpon"},
            {"url": "https://www.ello.co/{}", "name": "Ello"},
            {"url": "https://www.producthunt.com/@{}", "name": "Product Hunt"},
            {"url": "https://www.snapchat.com/add/{}", "name": "Snapchat"},
            {"url": "https://www.telegram.me/{}", "name": "Telegram"},
            {"url": "https://www.weheartit.com/{}", "name": "We Heart It"}
        ]

    def track(self):
        try:
            for site in self.social_media:
                url = site['url'].format(self.username)
                response = requests.get(url)
                if response.status_code == 200:
                    self.results[site['name']] = url
                else:
                    self.results[site['name']] = "Username not found!"
        except requests.RequestException as e:
            print(f"Ошибка при подключении к {site['name']}: {e}")

    def show_results(self):
        print(Fore.GREEN + f"\nРезультаты поиска для {self.username}:")
        for site, result in self.results.items():
            print(f"{site}: {result}")

class PhoneNumberInfo:
    def __init__(self):
        # Список API-ключей
        self.api_keys = [
            "83f567f1d514285201c48ce6669beb16",
            "c4a963e65cc53da7faaf0d2d39ae2955",
            "43637603eda5160dfcd8c8c5ad25cbe7",
            "0ad04fb3a06a51cc86429d6d9d4bf1d2",
            "90c90131b6b0ad8eb69b95ddc62f8968",
            "3e217160a7e61ba9ab9cb50fafa23a2c",
            "be3e4e957020132d631215e433c058ff",
            "1849da3c2ebf926fc4720d17797108ca",
            "9b2063919453d3f4caef5d32f128ffdf",
            "1e55b812d95d85b2688f1ecfb4ce0265",
            "bf0fdc3d48531aec6eff34d865896faa",
            "3017c4832557c5550f8080f374c29d14",
            "a6532784f3cc688179004418d9a4159a",
            "911bd53f1bded96ebcbead9225b96af1",
            "bacc0fffe4c1517864de86e1447a04e8"
            # Добавьте дополнительные ключи, если нужно
        ]
        # Шаблон для URL API с форматом json, номером телефона и ключом
        self.htmlweb_api_url = "http://htmlweb.ru/geo/api.php?json&telcod={user_number}&api_key={api_key}"
        self.cache_file = 'infa_phone.json'
        self._load_cache()
        self.request_limit = 5  # Лимит на количество запросов в день

    def _load_cache(self):
        try:
            with open(self.cache_file, 'r') as file:
                self.cache = json.load(file)
        except FileNotFoundError:
            self.cache = {}

        # Инициализируем счетчик запросов, если его нет
        if 'daily_requests' not in self.cache:
            self.cache['daily_requests'] = {
                'date': str(datetime.now().date()),
                'count': 0
            }

    def _save_cache(self):
        with open(self.cache_file, 'w') as file:
            json.dump(self.cache, file)

    def _check_request_limit(self):
        today = str(datetime.now().date())
        if self.cache['daily_requests']['date'] != today:
            # Если дата изменилась, сбросить счетчик запросов
            self.cache['daily_requests'] = {'date': today, 'count': 0}

        remaining_requests = self.request_limit - self.cache['daily_requests']['count']
        if remaining_requests > 0:
            return True, remaining_requests
        else:
            return False, remaining_requests

    def _increment_request_count(self):
        self.cache['daily_requests']['count'] += 1
        self._save_cache()

    def get_number_data(self, user_number):
        can_request, remaining_requests = self._check_request_limit()
        if not can_request:
            return {"status_error": True, "message": "Превышен дневной лимит запросов.", "remaining_requests": remaining_requests}

        if user_number in self.cache:
            return self.cache[user_number]

        # Выбираем случайный API-ключ
        api_key = random.choice(self.api_keys)
        # Формируем URL запроса с номером телефона и API-ключом
        api_url = self.htmlweb_api_url.format(user_number=user_number, api_key=api_key)
        response_htmlweb = requests.get(api_url)
        if response_htmlweb.ok:
            data_htmlweb = response_htmlweb.json()

            self.cache[user_number] = data_htmlweb
            self._increment_request_count()
            return data_htmlweb
        else:
            return {"status_error": True}

    def print_number_info(self):
        user_number = input(Fore.GREEN + "Введите номер телефона '+': ").strip()

        if user_number:
            print("Поиск данных...\n")
            data = self.get_number_data(user_number)
            if data.get("status_error"):
                if "message" in data:
                    print(Fore.RED + data["message"])
                else:
                    print(Fore.RED + "Ошибка: Не удалось получить данные. Проверьте номер телефона и попробуйте снова.")
                print(Fore.YELLOW + f"Оставшиеся запросы на сегодня: {data.get('remaining_requests', 0)}")
                return

            if data.get("limit") == 0:
                print(Fore.GREEN + "Вы израсходовали все лимиты запросов. Включите VPN или подождите 24ч")
                return

            country = data.get('country', {})
            region = data.get('region', {})
            other = data.get('0', {})

            print(Fore.BLUE + f"Страна: {country.get('name', 'Не найдено')}, {country.get('fullname', 'Не найдено')}")
            print(Fore.BLUE + f"Город: {other.get('name', 'Не найдено')}")
            print(Fore.BLUE + f"Почтовый индекс: {other.get('post', 'Не найдено')}")
            print(Fore.BLUE + f"Код валюты: {country.get('iso', 'Не найдено')}")
            print(Fore.BLUE + f"Телефонные коды: {data.get('capital', {}).get('telcod', 'Не найдено')}")
            print(Fore.BLUE + f"Посмотреть в wiki: {other.get('wiki', 'Не найдено')}")
            print(Fore.BLUE + f"Гос. номер региона авто: {region.get('autocod', 'Не найдено')}")
            print(Fore.BLUE + f"Оператор: {other.get('oper', 'Не найдено')}, {other.get('oper_brand', 'Не найдено')}, {other.get('def', 'Не найдено')}")
            print(Fore.BLUE + f"Местоположение: {country.get('name', 'Не найдено')}, {region.get('name', 'Не найдено')}, {other.get('name', 'Не найдено')} ({region.get('okrug', 'Не найдено')})")

            latitude = other.get('latitude', 'Не найдено')
            longitude = other.get('longitude', 'Не найдено')
            location = data.get('location', 'Не найдено')
            lang = country.get('lang', 'Не найдено').title()
            lang_code = country.get('langcod', 'Не найдено')
            capital = data.get('capital', {}).get('name', 'Не найдено')

            print(Fore.BLUE + f"Открыть на карте (google): https://www.google.com/maps/place/{latitude}+{longitude}")
            print(Fore.BLUE + f"Локация: {location}")
            print(Fore.BLUE + f"Язык общения: {lang}, {lang_code}")
            print(Fore.BLUE + f"Край/Округ/Область: {region.get('name', 'Не найдено')}, {region.get('okrug', 'Не найдено')}")
            print(Fore.BLUE + f"Столица: {capital}")
            print(Fore.BLUE + f"Широта/Долгота: {latitude}, {longitude}")
            print(Fore.BLUE + f"Оценка номера в сети: https://phoneradar.ru/phone/{user_number}")
            print(Fore.RED + " ")
            print(Fore.RED  + f"Whatsapp : https://wa.me/{user_number}")
            print(Fore.RED +  f"Viber -> https://viber.click/{user_number}")
            print(Fore.RED + f"Telegram -> https://t.me/{user_number}")
            # Вывод оставшегося количества запросов
            remaining_requests = self.request_limit - self.cache['daily_requests']['count']
            print(Fore.YELLOW + f"Оставшиеся запросы на сегодня: {remaining_requests}")
        else:
            print(Fore.RED + "Ошибка: Номер телефона не введен.")





class ExcelSearchertg:
    def __init__(self, folder_path, max_workers=4):
        self.folder_path = folder_path
        self.max_workers = max_workers  # Количество потоков для параллельной обработки
    
    def search_in_file(self, file_path, search_term):
        results = []
        try:
            # Загрузка Excel-файла с помощью openpyxl
            workbook = load_workbook(file_path, data_only=True)
            
            # Перебор всех листов
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Получаем заголовки столбцов
                headers = [cell.value for cell in sheet[1]]  # Предполагаем, что заголовки находятся в первой строке
                
                # Перебор всех строк на листе
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if any(str(cell) == search_term for cell in row if cell is not None):
                        # Формирование результата при нахождении совпадения
                        matched_data = {headers[i] if i < len(headers) else f"Column {i+1}": cell for i, cell in enumerate(row)}
                        result = {
                            "file": os.path.basename(file_path),
                            "sheet": sheet_name,
                            "data": matched_data
                        }
                        results.append(result)
                        break  # Останавливаем обработку после первого найденного совпадения
                
        except Exception as e:
            print(f"Ошибка при обработке файла {file_path}: {e}")
        
        return results

    def format_output(self, result, search_term):
        print(f"Число '{search_term}' найдено в файле '{result['file']}', лист '{result['sheet']}':\n")
        for column_name, value in result['data'].items():
            # Учитываем, что значение может быть числом, строкой или чем-то другим
            formatted_value = value if isinstance(value, (int, float)) else str(value)
            print(f"{column_name} - {formatted_value}")
        print("\n" + "="*40 + "\n")  # Разделение между записями

    def search_in_folder(self, search_term):
        results = []
        # Список всех файлов .xlsx в папке
        files = [os.path.join(self.folder_path, f) for f in os.listdir(self.folder_path) if f.endswith(".xlsx")]
        
        # Использование многопоточности для параллельной обработки файлов
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # Создаем словарь для хранения будущих результатов
            future_to_file = {executor.submit(self.search_in_file, file, search_term): file for file in files}
            
            # Используем tqdm для отображения прогресса
            for future in tqdm(as_completed(future_to_file), total=len(future_to_file), desc="Поиск"):
                try:
                    file_results = future.result()
                    results.extend(file_results)
                except Exception as e:
                    print(f"Ошибка в процессе обработки: {e}")

        # Вывод результатов
        for result in results:
            self.format_output(result, search_term)


class ExcelSearcherphone:
    def __init__(self, folder_path, max_workers=4):
        self.folder_path = folder_path
        self.max_workers = max_workers  # Количество потоков для параллельной обработки
    
    def search_in_file(self, file_path, search_term):
        results = []
        try:
            # Загрузка Excel-файла с помощью openpyxl
            workbook = load_workbook(file_path, data_only=True)
            
            # Перебор всех листов
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Получаем заголовки столбцов
                headers = [cell.value for cell in sheet[1]]  # Предполагаем, что заголовки находятся в первой строке
                
                # Перебор всех строк на листе
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if any(str(cell) == search_term for cell in row if cell is not None):
                        # Формирование результата при нахождении совпадения
                        matched_data = {headers[i] if i < len(headers) else f"Column {i+1}": cell for i, cell in enumerate(row)}
                        result = {
                            "file": os.path.basename(file_path),
                            "sheet": sheet_name,
                            "data": matched_data
                        }
                        results.append(result)
                        break  # Останавливаем обработку после первого найденного совпадения
                
        except Exception as e:
            print(f"Ошибка при обработке файла {file_path}: {e}")
        
        return results

    def format_output(self, result, search_term):
        print(f"Число '{search_term}' найдено в файле '{result['file']}', лист '{result['sheet']}':\n")
        for column_name, value in result['data'].items():
            # Учитываем, что значение может быть числом, строкой или чем-то другим
            formatted_value = value if isinstance(value, (int, float)) else str(value)
            print(f"{column_name} - {formatted_value}")
        print("\n" + "="*40 + "\n")  # Разделение между записями

    def search_in_folder(self, search_term):
        results = []
        # Список всех файлов .xlsx в папке
        files = [os.path.join(self.folder_path, f) for f in os.listdir(self.folder_path) if f.endswith(".xlsx")]
        
        # Использование многопоточности для параллельной обработки файлов
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # Создаем словарь для хранения будущих результатов
            future_to_file = {executor.submit(self.search_in_file, file, search_term): file for file in files}
            
            # Используем tqdm для отображения прогресса
            for future in tqdm(as_completed(future_to_file), total=len(future_to_file), desc="Поиск"):
                try:
                    file_results = future.result()
                    results.extend(file_results)
                except Exception as e:
                    print(f"Ошибка в процессе обработки: {e}")

        # Вывод результатов
        for result in results:
            self.format_output(result, search_term)

if __name__ == "__main__":
    main()


